import csv
from io import TextIOWrapper, StringIO, BytesIO

# librerias e importaciones para excel
import io
from flask import request, jsonify, send_file, current_app
from openpyxl import load_workbook
from openpyxl.styles import Alignment
# Importacion del link OneDrive MANTENIEMIENTO desde el archivo config
from config import LinkOneDriveMantenimiento
# Importacion del link OneDrive CALIBRACION desde el archivo config
from config import LinkOneDriveCalibracion

from flask import Flask, render_template, request, redirect, url_for, flash, Response
from flask_mysqldb import MySQL,MySQLdb
from flask_wtf.csrf import CSRFProtect
from flask_login import LoginManager, login_user, logout_user, login_required
from config import config
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta

# Para subir archivos tipo foto al servidor
import os
import shutil
from werkzeug.utils import secure_filename

# Models:
from models.ModelUser import ModelUser

# Entities:
from models.entities.User import User
# Para el modulo json que se esta utilizan para el checked
from flask import Blueprint, Flask, render_template, request, jsonify 
# Importaciones para el envio de correos en prestamo
from flask_login import current_user
# Importaciones desde el archivo email_service
from email_service import send_email_with_logo
from email_service import send_prestamo_notification_html
# Importaciones desde el archivo email_devolucion
from email_devolucion import send_email_envio_with_logo
from email_devolucion import send_devolucion_notification_html

from extensions import db, login_manager


bp = Blueprint('inventario', __name__)

@bp.context_processor
def link_onedrive_mantenimiento():
    return dict(onedrive_link_mantenimiento=LinkOneDriveMantenimiento.ONEDRIVE_LINK_MANTENIMIENTO)

@bp.context_processor
def link_onedrive_calibracion():
    return dict(onedrive_link_calibracion=LinkOneDriveCalibracion.ONEDRIVE_LINK_CALIBRACION)

@login_manager.user_loader
def load_user(id):
    return ModelUser.get_by_id(db, id)

@bp.after_request
def evita_cache(response):
    response.cache_control.no_store = True
    response.cache_control.no_cache = True
    response.cache_control.must_revalidate = True
    response.cache_control.max_age = 0
    response.expires = 0
    response.pragma = 'no-cache'
    return response

@bp.route('/')
# @login_required
def index():
    return redirect(url_for('inventario.login'))


@bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User(0, request.form['username'], request.form['password'])
        logged_user = ModelUser.login(db, user)

        if logged_user:
            if logged_user.password:
                login_user(logged_user)

                if logged_user.rol in ['salud', 'gastronomia', 'lacma', 'arquitectura', 'tecnologia']:
                    return redirect(url_for('inventario.index_modulo', modulo=logged_user.rol))
                elif logged_user.rol == 'admin':
                    return redirect(url_for('inventario.home'))
                else:
                    flash('Rol no autorizado')
                    return redirect(url_for('login'))
            else:
                flash("Contraseña incorrecta...")
        else:
            flash("Usuario no encontrado...")
    return render_template('auth/login.html')


@bp.route('/logout')
# @login_required
def logout():
    logout_user()
    return redirect(url_for('inventario.login'))

@bp.route('/home')
@login_required
def home():
    cur = db.connection.cursor()

    # Obtener la fecha actual
    fecha_actual = datetime.now().date()

    # Consultar la cantidad de equipos según el vencimiento del mantenimiento
    cur.execute("""
        SELECT 
            SUM(DATEDIFF(vencimiento_mantenimiento, %s) < 0) AS vencidas,
            SUM(DATEDIFF(vencimiento_mantenimiento, %s) BETWEEN 0 AND 30) AS proximas,
            SUM(DATEDIFF(vencimiento_mantenimiento, %s) BETWEEN 31 AND 90) AS entre_31_y_90,
            SUM(DATEDIFF(vencimiento_mantenimiento, %s) > 90) AS mas_90_dias
        FROM indexssalud WHERE enable = 1 and de_baja = 0
    """, (fecha_actual, fecha_actual, fecha_actual, fecha_actual))

    # Obtener los resultados
    resultados_mantenimiento = cur.fetchone()
    vencidas_mantenimiento = resultados_mantenimiento[0]
    proximas_mantenimiento = resultados_mantenimiento[1]
    entre_31_y_90_mantenimiento = resultados_mantenimiento[2]
    mas_90_dias_mantenimiento = resultados_mantenimiento[3]

    print(resultados_mantenimiento)
   

    # Consultar la cantidad de equipos según el vencimiento de la calibración
    cur.execute("""
        SELECT 
            SUM(DATEDIFF(vencimiento_calibracion, %s) < 0) AS vencidas,
            SUM(DATEDIFF(vencimiento_calibracion, %s) BETWEEN 0 AND 30) AS proximas,
            SUM(DATEDIFF(vencimiento_calibracion, %s) BETWEEN 31 AND 90) AS entre_31_y_90,
            SUM(DATEDIFF(vencimiento_calibracion, %s) > 90) AS mas_90_dias
        FROM indexssalud WHERE enable = 1 and de_baja = 0
    """, (fecha_actual, fecha_actual, fecha_actual, fecha_actual))

    # Obtener los resultados del vencimiento de calibración
    resultados_calibracion = cur.fetchone()
    vencidas_calibracion = resultados_calibracion[0]
    proximas_calibracion = resultados_calibracion[1]
    entre_31_y_90_calibracion = resultados_calibracion[2]
    mas_90_dias_calibracion = resultados_calibracion[3]

    return render_template('home.html', vencidas_mantenimiento=vencidas_mantenimiento, 
                                        proximas_mantenimiento=proximas_mantenimiento, 
                                        entre_31_y_90_mantenimiento=entre_31_y_90_mantenimiento, 
                                        mas_90_dias_mantenimiento=mas_90_dias_mantenimiento,
                                        vencidas_calibracion=vencidas_calibracion, 
                                        proximas_calibracion=proximas_calibracion, 
                                        entre_31_y_90_calibracion=entre_31_y_90_calibracion, 
                                        mas_90_dias_calibracion=mas_90_dias_calibracion)



# --------------------------- INICIA MODULO DE SALUD-----------------------------
# --------------------------- DATOS PROVEEDOR SALUD --------------------------------
@bp.route('/datosProveedorSalud/<id>')
@login_required
def datosProveedorSalud(id):
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    # cur = db.connection.cursor()
    cur.execute('SELECT * FROM datosproveedorsalud WHERE id = %s', [id])
    proveedor = cur.fetchall()
    print (proveedor)
    return render_template('datosProveedorSalud.html', datosproveedorsalud=proveedor)

# FUNCIÓN ACTUALIZAR DATOS PROVEEDOR
@bp.route('/update_datos_proveedor_salud/<id>', methods = ['POST'])
def ACTUALIZAR_DATOS_PROVEEDOR_SALUD(id):
    if request.method =='POST':
        telefono_empresa = request.form ['telefono_empresa']
        nombre_empresa = request.form ['nombre_empresa']
        nombre_contacto = request.form ['nombre_contacto']
        correo = request.form ['correo']
        cargo_contacto = request.form ['cargo_contacto']
        ciudad = request.form ['ciudad']
        cur = db.connection.cursor() 
        cur.execute(""" UPDATE datosproveedorsalud SET telefono_empresa = %s, nombre_empresa = %s, nombre_contacto = %s, correo = %s, cargo_contacto = %s, ciudad = %s WHERE id = %s """, 
                                                      (telefono_empresa, nombre_empresa, nombre_contacto, correo, cargo_contacto, ciudad, id))
        db.connection.commit()
    flash('Datos actualizados satisfactorimanete', 'success')
    return redirect(url_for('inventario.datosProveedorSalud', id = id))

# ESTA FUNCIÓN ME LLEVA A OTRA VISTA PARA AGREGAR LOS NUEVOS PROVEEDORES
@bp.route('/agregarNuevoProveedor')
@login_required
def agregarNuevoProveedor():
    return render_template('agregarNuevoProveedor.html')

@bp.route('/add_datos_proveedor_salud', methods=['POST'])
def AGREGAR_DATOS_PROVEEDOR_SALUD():
    if request.method =='POST':
        telefono_empresa = request.form ['telefono_empresa']
        nombre_empresa = request.form ['nombre_empresa']
        nombre_contacto = request.form ['nombre_contacto']
        correo = request.form ['correo']
        cargo_contacto = request.form ['cargo_contacto']
        ciudad = request.form ['ciudad']

         # Valida que todos los campos estén diligenciados
        if not telefono_empresa or not nombre_empresa or not nombre_contacto or not ciudad:
            flash('Todos los campos son obligatorios', 'danger')
            return redirect(url_for('agregarNuevoProveedor'))
        
        cur = db.connection.cursor() 
        cur.execute('INSERT INTO datosproveedorsalud (telefono_empresa, nombre_empresa, nombre_contacto, correo, cargo_contacto, ciudad) VALUES (%s, %s, %s, %s, %s, %s)', 
                                                     (telefono_empresa, nombre_empresa, nombre_contacto, correo, cargo_contacto, ciudad))
        db.connection.commit()
    flash ('Datos agregados satisfactoriamente', 'success')
    return redirect(url_for('inventario.agregarNuevoProveedor')) 

# ---------------------------FUNCIÓN PARA EL MANEJO DE LOS MODULOS-----------------------------
@bp.route('/<modulo>')
@login_required
def index_modulo(modulo):
    modulos_validos = ['salud', 'gastronomia', 'lacma', 'arquitectura']
    
    if modulo not in modulos_validos:
        # flash("Modulo no válido", "error")
        return redirect(url_for('inventario.home'))  # <-- redirige al home si no existe
    
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

    # Traer equipos solo del modulo actual
    cur.execute("""SELECT i.*, p.enable_prestamos FROM indexssalud i LEFT JOIN prestamos_equiposalud p ON i.cod_articulo = p.cod_articulo AND p.enable_prestamos = 1 WHERE i.enable=1 AND i.de_baja=0 AND i.modulo=%s""", (modulo,))
    equipos = cur.fetchall()

    # Traer proveedores, estados y ubicaciones
    cur.execute('SELECT id, nombre_empresa FROM datosproveedorsalud')
    proveedores = cur.fetchall()

    cur.execute('SELECT id, estado_equipo FROM estados_equipos')
    estadoEquipos = cur.fetchall()

    cur.execute('SELECT id, ubicacion_original FROM ubicacion_equipos')
    ubicacionEquipos = cur.fetchall()

    return render_template(f'indexSalud.html', indexssalud=equipos, proveedores=proveedores, estadoEquipos=estadoEquipos, ubicacionEquipos=ubicacionEquipos, modulo=modulo)
# --------------------------- -----------------------------
# @app.route('/indexSalud')
# @login_required
# def indexSalud():
#     cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
#     # cur = db.connection.cursor()
#     cur.execute('SELECT i.*, p.enable_prestamos FROM indexssalud i LEFT JOIN prestamos_equiposalud p ON i.cod_articulo = p.cod_articulo AND p.enable_prestamos = 1 WHERE i.enable=1 AND i.de_baja=0') #A raiz del enable=1 no se deben eliminar en la DB
#     data = cur.fetchall()

#     cur.execute('SELECT id, nombre_empresa FROM datosproveedorsalud')
#     proveedores = cur.fetchall()

#     cur.execute('SELECT id, estado_equipo FROM estados_equipos')
#     estadoEquipos = cur.fetchall()

#     cur.execute('SELECT id, ubicacion_original FROM ubicacion_equipos')
#     ubicacionEquipos = cur.fetchall()
#     # print(ubicacionEquipos)
#     return render_template('indexSalud.html', indexssalud=data, proveedores=proveedores, estadoEquipos=estadoEquipos, ubicacionEquipos=ubicacionEquipos)

@bp.route('/<modulo>/add_productoSalud', methods=['POST'])
def AGREGAR_PRODUCTO_SALUD(modulo):
    # Validar que el modulo exista
    # if modulo not in ['salud', 'gastronomia', 'lacma', 'arquitectura']:
    #     flash("Modulo no válido", "error")
    #     return redirect(url_for('home'))

    if request.method =='POST':
        cod_articulo = request.form ['cod_articulo']
        nombre_equipo = request.form ['nombre_equipo']

        # Validación de cod_articulo
        try:
            cod_articulo = int(cod_articulo)
        except ValueError:
            flash('Por favor ingresar solo números en el código del equipo', 'error')
            return redirect(url_for('index_modulo', modulo=modulo))

        # Consulta para verificar si el cod_articulo ya existe en la base de datos
        cur = db.connection.cursor()
        cur.execute("SELECT * FROM indexssalud WHERE cod_articulo = %s", (cod_articulo,))
        existing_articulo = cur.fetchone()

        if existing_articulo:
            flash(f'El código de equipo {cod_articulo} ya existe', 'error')
            return redirect(url_for('inventario.index_modulo', modulo=modulo))

        # PARA EL CHECKBOX Y SEMAFORO DE MANTENIMIENTO
        fecha_mantenimiento = request.form ['fecha_mantenimiento']
        vencimiento_mantenimiento = request.form ['vencimiento_mantenimiento']
        checkbox_mantenimiento = 'Inactivo' # Valor predeterminado
        
        # # Obtener hora actual del equipo
        # hora_actual = datetime.now().date()
        color = 'verde'

        # if  vencimiento_mantenimiento:

        #     vencimiento_mant = datetime.strptime(vencimiento_mantenimiento, '%Y-%m-%d').date()
        #     if vencimiento_mant < hora_actual + timedelta(days=0):
        #         color = 'purple'  # Falta menos de un mes
        #     elif vencimiento_mant <= hora_actual + timedelta(days=30):
        #         color = 'red'  # Falta menos de tres meses
        #     elif vencimiento_mant <= hora_actual + timedelta(days=90):
        #         color = 'yellow'  # Falta menos de tres meses
        # else:
        #     flash('Debe ingresar las fechas de mantenimiento.', 'error')
        #     return redirect(url_for('indexSalud'))
        if not vencimiento_mantenimiento:
            flash('Debe ingresar la fecha vencimiento de mantenimiento.', 'error')
            print ("Color semaforo:", vencimiento_mantenimiento)
            return redirect(url_for('index_modulo', modulo=modulo))
        

        # PARA EL CHECKBOX DE CALIBRACIÓN
        fecha_calibracion = request.form ['fecha_calibracion'] or None
        vencimiento_calibracion = request.form ['vencimiento_calibracion'] or None
        checkbox_calibracion = 'Inactivo' # Valor predeterminado

        fecha_ingreso = request.form ['fecha_ingreso']
        # Validación de fecha_ingreso
        if not fecha_ingreso or fecha_ingreso.strip() == '':
            flash('Verifique la fecha de ingreso. El campo no puede estar vacío.', 'error')
            return redirect(url_for('inventario.index_modulo', modulo=modulo))
        
        periodicidad = request.form ['periodicidad']
        # Validación de periodicidad
        try:
            periodicidad = int (periodicidad)
        except ValueError:
            flash('Por favor ingresar solo números en la periodicidad del equipo', 'error')
            return redirect(url_for('inventario.index_modulo', modulo=modulo))
        
        estado_equipo = request.form ['estado_equipo']
        ubicacion_original = request.form ['ubicacion_original']
        garantia = request.form ['garantia']
        criticos = request.form ['criticos']
        proveedor_responsable = request.form ['proveedor_responsable']

        # Manejo de la imagen
        if 'imagen_producto' not in request.files:
            flash('No existe archivo de imagen.', 'error')
            return redirect(url_for('inventario.index_modulo', modulo=modulo))

        file = request.files['imagen_producto']

        if file.filename == '':
            flash('Por favor seleccione un archivo de imagen.', 'error')
            return redirect(url_for('inventario.index_modulo', modulo=modulo))

        # Validar extensión de imagen
        extensiones_permitidas_img = ('.png', '.jpg', '.jpeg')
        if not file.filename.lower().endswith(extensiones_permitidas_img):
            flash(f'Formato de imagen no permitido. Solo se permiten: {", ".join(extensiones_permitidas_img)}', 'error')
            return redirect(url_for('inventario.index_modulo', modulo=modulo))

        # Guardar imagen
        filename = secure_filename(file.filename)
        filepath_to_db_img = os.path.join('fotos', filename).replace("\\", "/")
        ruta_absoluta = os.path.join(bp.root_path, 'static', filepath_to_db_img)
        file.save(ruta_absoluta)

        # Manejo del PDF
        filepath_to_db_pdf = None
        if 'guia_pdf' in request.files:
            file_pdf = request.files['guia_pdf']
            if file_pdf and file_pdf.filename != '':
                # Validar extensión del PDF
                if not file_pdf.filename.lower().endswith('.pdf'):
                    flash('Formato no permitido. Solo se aceptan archivos PDF.', 'error')
                    return redirect(url_for('inventario.index_modulo', modulo=modulo))

        # Guardar PDF
        filename_pdf = secure_filename(file_pdf.filename)
        filepath_to_db_pdf = os.path.join('pdf', filename_pdf).replace("\\", "/")
        ruta_absoluta_pdf = os.path.join(bp.root_path, 'static', filepath_to_db_pdf)
        file_pdf.save(ruta_absoluta_pdf)

        especificaciones_instalacion = request.form ['especificaciones_instalacion']
        cuidados_basicos = request.form ['cuidados_basicos']
        periodicidad_calibracion = request.form ['periodicidad_calibracion']
        marca_equipo_salud = request.form ['marca_equipo_salud']
        modelo_equipo_salud = request.form ['modelo_equipo_salud']
        serial_equipo_salud = request.form ['serial_equipo_salud']
        fecha_de_baja = date.today() if estado_equipo == "DE BAJA" else None

        cur = db.connection.cursor()
        if estado_equipo == "DE BAJA":
            # Guardar en la tabla de equipos de baja
            cur.execute("""INSERT INTO equipossalud_debaja (cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, 
                                                            vencimiento_calibracion, fecha_ingreso, periodicidad, estado_equipo, ubicacion_original, garantia, criticos, proveedor_responsable, 
                                                            color, checkbox_mantenimiento, checkbox_calibracion, imagen, especificaciones_instalacion, cuidados_basicos, periodicidad_calibracion, 
                                                            marca_equipo_salud, modelo_equipo_salud, serial_equipo_salud, fecha_de_baja) VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (
                    cod_articulo,
                    nombre_equipo,
                    fecha_mantenimiento,
                    vencimiento_mantenimiento,
                    fecha_calibracion,
                    vencimiento_calibracion,
                    fecha_ingreso,
                    periodicidad,
                    estado_equipo,
                    ubicacion_original,
                    garantia,
                    criticos,
                    proveedor_responsable,
                    color,
                    checkbox_mantenimiento,
                    checkbox_calibracion,
                    filepath_to_db_img,
                    especificaciones_instalacion,
                    cuidados_basicos,
                    periodicidad_calibracion,
                    marca_equipo_salud,
                    modelo_equipo_salud,
                    serial_equipo_salud,
                    fecha_de_baja
                    
                ),
            )
        else:   
            print("Insertando equipo en modulo:", modulo)
            # Guardar en la tabla de equipos de indexssalud
            cur.execute("""INSERT INTO indexssalud (cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion, fecha_ingreso, periodicidad,
                           estado_equipo, ubicacion_original, garantia, criticos, proveedor_responsable, color, checkbox_mantenimiento, checkbox_calibracion, imagen, especificaciones_instalacion, cuidados_basicos,
                           periodicidad_calibracion, marca_equipo_salud, modelo_equipo_salud, serial_equipo_salud, pdf_salud, modulo) VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (
                    cod_articulo,
                    nombre_equipo,
                    fecha_mantenimiento,
                    vencimiento_mantenimiento,
                    fecha_calibracion,
                    vencimiento_calibracion,
                    fecha_ingreso,
                    periodicidad,
                    estado_equipo,
                    ubicacion_original,
                    garantia,
                    criticos,
                    proveedor_responsable,
                    color,
                    checkbox_mantenimiento,
                    checkbox_calibracion,
                    filepath_to_db_img,
                    especificaciones_instalacion,
                    cuidados_basicos,
                    periodicidad_calibracion,
                    marca_equipo_salud,
                    modelo_equipo_salud,
                    serial_equipo_salud,
                    filepath_to_db_pdf,
                    modulo

                ),
            )
        db.connection.commit()

        flash (f"Equipo agregado correctamente al módulo {modulo}", "success")
        return redirect(url_for('inventario.index_modulo', modulo=modulo)) 
    
# ---------------------------FUNCION PARA CARGAR IMAGEN DEL EQUIPO DESDE LA TABLA indexSalud EN EL CAMPO ACCIONES SUBIR_IMAGEN-----------------------------  
@bp.route('/subir_imagen/<int:id_producto>', methods=['POST'])
def subir_imagen(id_producto):
    if 'imagen_producto' not in request.files:
        flash('No se seleccionó ningún archivo', 'error')
        return redirect(url_for('index_modulo'))

    file = request.files['imagen_producto']
    if file.filename == '':
        flash('Por favor seleccione un archivo válido', 'error')
        return redirect(url_for('index_modulo'))
    
    # Validar extensión
    if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg')):
        flash('Solo se permiten archivos PNG, JPG', 'error')
        return redirect(url_for('index_modulo'))

    if file:
        filename = secure_filename(file.filename)
        filepath_to_db_img = os.path.join('fotos', filename).replace("\\", "/")
        ruta_absoluta = os.path.join(bp.root_path, 'static', filepath_to_db_img)

        # Guardar en disco
        file.save(ruta_absoluta)

        # Actualizar en BD
        cur = db.connection.cursor()
        cur.execute("""
            UPDATE indexssalud 
            SET imagen = %s 
            WHERE id = %s
        """, (filepath_to_db_img, id_producto))
        db.connection.commit()
        cur.close()

        flash('Imagen cargada correctamente', 'success')
        return redirect(url_for('inventario.index_modulo', modulo='modulo'))

# ---------------------------FUNCION PARA CARGAR PDFS DEL EQUIPO DESDE LA TABLA indexSalud EN EL CAMPO ACCIONES SUBIR_GUIA---------------------------- 
@bp.route('/subir_pdf/<int:id_producto>', methods=['GET', 'POST'])
def subir_pdf(id_producto):

    # ---- GET ----
    if request.method == 'GET':
        flash("Debe seleccionar un archivo PDF.", "warning")
        return redirect(url_for('inventario.index_modulo', modulo="salud"))

    # ---- POST ----
    if 'pdf_salud' not in request.files:
        flash('No se seleccionó ningún archivo', 'error')
        return redirect(url_for('inventario.index_modulo', modulo="salud"))

    file = request.files['pdf_salud']
    if file.filename == '':
        flash('Por favor seleccione un archivo válido', 'error')
        return redirect(url_for('inventario.index_modulo', modulo="salud"))

    if not file.filename.lower().endswith('.pdf'):
        flash('Solo se permiten archivos PDF', 'error')
        return redirect(url_for('inventario.index_modulo', modulo="salud"))

    # Guardar archivo
    filename = secure_filename(file.filename)
    filepath_to_db_pdf = os.path.join('pdf', filename).replace("\\", "/")
    ruta_absoluta = os.path.join(bp.root_path, 'static', filepath_to_db_pdf)
    file.save(ruta_absoluta)

    cur = db.connection.cursor()
    cur.execute("""
        UPDATE indexssalud 
        SET pdf_salud = %s 
        WHERE id = %s
    """, (filepath_to_db_pdf, id_producto))
    db.connection.commit()
    cur.close()

    flash('Guía cargada correctamente', 'success')
    return redirect(url_for('inventario.index_modulo', modulo="modulo"))


# ---------------------------INICIA INSERT MASIVO DE EQUIPOS CSV DE SALUD-----------------------------
# ---------------------------INSERT MASIVO DE EQUIPOS CSV DE SALUD----------------------
@bp.route('/<modulo>/insert_csv', methods=['POST'])
@login_required
def insert_csv(modulo):
    file = request.files.get('file')
    if not file or file.filename == '':
        flash('No seleccionó ningún archivo', 'error')
        return redirect(url_for('inventario.index_modulo', modulo=modulo))

    # Cursor normal
    cur = db.connection.cursor()

    # 1. Cargar todos los proveedores UNA sola vez y crear el mapa nombre -> id
    cur.execute("SELECT id, nombre_empresa FROM datosproveedorsalud")
    proveedores = cur.fetchall()
    proveedor_map = {p[1].strip().lower(): p[0] for p in proveedores}

    # 2. Preparar lectura del CSV
    file = TextIOWrapper(file, encoding='latin-1')
    csv_reader = csv.reader(file)
    next(csv_reader, None)  # Saltar encabezado si existe

    fotos_folder = os.path.join(os.path.dirname(__file__), 'static', 'fotos')

    codigos_duplicados = []
    datos_validos = []

    # 3. Primera pasada: validar filas y recolectar datos válidos
    for row in csv_reader:
        if not row:
            continue

        cod_articulo_str = row[0]

        # Validar cod_articulo numérico
        try:
            cod_articulo = int(cod_articulo_str)
        except (ValueError, TypeError):
            flash(f'Código inválido: {cod_articulo_str}', 'error')
            continue

        # Verificar duplicados en indexssalud
        cur.execute("SELECT cod_articulo FROM indexssalud WHERE cod_articulo = %s", (cod_articulo,))
        if cur.fetchone():
            codigos_duplicados.append(str(cod_articulo))
            continue

        # Verificar duplicados en equipossalud_debaja
        cur.execute("SELECT cod_articulo FROM equipossalud_debaja WHERE cod_articulo = %s", (cod_articulo,))
        if cur.fetchone():
            codigos_duplicados.append(str(cod_articulo))
            continue

        # Verificar que la imagen exista
        imagen = row[14]
        imagen_path = os.path.join(fotos_folder, imagen)
        if not os.path.isfile(imagen_path):
            flash(f'Imagen no encontrada: {imagen}', 'error')
            continue

        # Validaciones básicas
        if not row[7]:
            flash(f'Equipo con código {cod_articulo} no tiene Mantenimiento Actual.', 'error')
            continue

        if not row[8]:
            flash(f'Equipo con código {cod_articulo} no tiene Vencimiento Mantenimiento.', 'error')
            continue

        if not row[9]:
            flash(f'Equipo con código {cod_articulo} no tiene periodicidad de calibración (si no tiene, ingresa 0).', 'error')
            continue

        datos_validos.append(row)

    # 4. Segunda pasada: insertar solo los datos válidos
    for row in datos_validos:
        cod_articulo = int(row[0])
        nombre_equipo = row[1]
        ubicacion_original = row[2]
        estado_equipo = row[3]
        fecha_ingreso = row[4]
        garantia = row[5]
        periodicidad = int(row[6])
        fecha_mantenimiento = row[7]
        vencimiento_mantenimiento = row[8]
        periodicidad_calibracion = int(row[9])
        fecha_calibracion = row[10] or None
        vencimiento_calibracion = row[11] or None
        criticos = row[12]

        nombre_proveedor = row[13].strip().lower()
        proveedor_responsable = proveedor_map.get(nombre_proveedor)

        if not proveedor_responsable:
            flash(f"Proveedor '{row[13]}' no encontrado en la base de datos.", 'error')
            continue

        imagen = row[14]
        especificaciones_instalacion = row[15]
        cuidados_basicos = row[16]
        marca_equipo_salud = row[17]
        modelo_equipo_salud = row[18]
        serial_equipo_salud = row[19]

        ruta_imagen_db = f'fotos/{secure_filename(imagen)}'
        checkbox_mantenimiento = 'Inactivo'
        checkbox_calibracion = 'Inactivo'
        fecha_de_baja = date.today() if estado_equipo == "DE BAJA" else None
        color = 'verde'

        if estado_equipo == 'DE BAJA':

            # 1️⃣ INSERTAR EN equipossalud_debaja
            cur.execute("""
                INSERT INTO equipossalud_debaja (
                    cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento,
                    fecha_calibracion, vencimiento_calibracion, fecha_ingreso, periodicidad,
                    estado_equipo, ubicacion_original, garantia, criticos, proveedor_responsable,
                    imagen, especificaciones_instalacion, cuidados_basicos,
                    periodicidad_calibracion, marca_equipo_salud, modelo_equipo_salud,
                    serial_equipo_salud, color, checkbox_mantenimiento, checkbox_calibracion,
                    fecha_de_baja, modulo
                ) VALUES (
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, 
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s
                )
            """, (
                cod_articulo, #ORGANIZAR BIEN LA FUNCIÓN 
                nombre_equipo,
                fecha_mantenimiento,
                vencimiento_mantenimiento,
                fecha_calibracion,
                vencimiento_calibracion,
                fecha_ingreso,
                periodicidad,
                estado_equipo,
                ubicacion_original,
                garantia,
                criticos,
                proveedor_responsable,
                ruta_imagen_db,
                especificaciones_instalacion,
                cuidados_basicos,
                periodicidad_calibracion,
                marca_equipo_salud,
                modelo_equipo_salud,
                serial_equipo_salud,
                color,
                checkbox_mantenimiento,
                checkbox_calibracion,
                fecha_de_baja,
                modulo
            ))

            # 2️⃣ INSERTAR TAMBIÉN EN indexssalud como DE BAJA (enable=0, de_baja=1)
            cur.execute("""
                INSERT INTO indexssalud (
                    cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento,
                    fecha_calibracion, vencimiento_calibracion, fecha_ingreso, periodicidad,
                    estado_equipo, ubicacion_original, garantia, criticos, proveedor_responsable,
                    imagen, especificaciones_instalacion, cuidados_basicos,
                    periodicidad_calibracion, marca_equipo_salud, modelo_equipo_salud,
                    serial_equipo_salud, color, checkbox_mantenimiento, checkbox_calibracion,
                    fecha_de_baja, enable, de_baja, modulo
                ) VALUES (
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s
                )
            """, (
                cod_articulo,
                nombre_equipo,
                fecha_mantenimiento,
                vencimiento_mantenimiento,
                fecha_calibracion,
                vencimiento_calibracion,
                fecha_ingreso,
                periodicidad,
                estado_equipo,              # DE BAJA
                ubicacion_original,
                garantia,
                criticos,
                proveedor_responsable,
                ruta_imagen_db,
                especificaciones_instalacion,
                cuidados_basicos,
                periodicidad_calibracion,
                marca_equipo_salud,
                modelo_equipo_salud,
                serial_equipo_salud,
                color,
                checkbox_mantenimiento,
                checkbox_calibracion,
                fecha_de_baja,
                0,      # enable = 0
                1,      # de_baja = 1
                modulo
            ))
        
        else:
            # Insertar estados activos en indexssalud
            cur.execute("""
                INSERT INTO indexssalud (
                    cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento,
                    fecha_calibracion, vencimiento_calibracion, fecha_ingreso, periodicidad,
                    estado_equipo, ubicacion_original, garantia, criticos, proveedor_responsable,
                    imagen, especificaciones_instalacion, cuidados_basicos,
                    periodicidad_calibracion, marca_equipo_salud, modelo_equipo_salud,
                    serial_equipo_salud, color, checkbox_mantenimiento, checkbox_calibracion,
                    fecha_de_baja, enable, de_baja, modulo
                ) VALUES (
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s
                )
            """, (
                cod_articulo,
                nombre_equipo,
                fecha_mantenimiento,
                vencimiento_mantenimiento,
                fecha_calibracion,
                vencimiento_calibracion,
                fecha_ingreso,
                periodicidad,
                estado_equipo,              # USO / MANTENIMIENTO / CALIBRACIÓN
                ubicacion_original,
                garantia,
                criticos,
                proveedor_responsable,
                ruta_imagen_db,
                especificaciones_instalacion,
                cuidados_basicos,
                periodicidad_calibracion,
                marca_equipo_salud,
                modelo_equipo_salud,
                serial_equipo_salud,
                color,
                checkbox_mantenimiento,
                checkbox_calibracion,
                None,   # fecha_de_baja NULL
                1,      # enable = 1
                0,      # de_baja = 0
                modulo
            ))

    db.connection.commit()

    if codigos_duplicados:
        flash(f'Los siguientes códigos ya existen y no fueron insertados: {", ".join(codigos_duplicados)}', 'error')
    if datos_validos:
        flash(f'{len(datos_validos)} equipos importados exitosamente.', 'success')
    else:
        flash('No se importó ningún equipo.', 'error')

    # Redirigir al módulo correcto (no al literal 'modulo')
    return redirect(url_for('inventario.index_modulo', modulo=modulo))

# ---------------------------FINALIZA INSERT MASIVO CSV DE SALUD-----------------------------

# ---------------------------INICIA ACTUALIZACIÓN MASIVA DE FECHAS DE MANTENIMIENTO, CALIBRACIÓN Y PERIODICIDAD EQUIPOS CSV DE SALUD-----------------------------
@bp.route('/updateDate_csv', methods=['POST'])
def updateDate_csv(modulo):
    if request.method == 'POST':
        file = request.files['file']

        if not file:
            flash('No selecciono ningun archivo')
            return redirect(url_for('index_modulo', modulo=modulo))
        
        # Decodifica la fila
        file = TextIOWrapper(file, encoding='latin-1')
        csv_reader = csv.reader(file)
        
        # Salta el encabezado
        next(csv_reader)
        
        # Leer cada fila y actualizar las fechas
        for row in csv_reader:
            cod_articulo = row[0]
            nombre_equipo = row[1]
            periodicidad = row[2]
            fecha_mantenimiento = datetime.strptime(row[3], '%Y/%m/%d')
            vencimiento_mantenimiento = datetime.strptime(row[4], '%Y/%m/%d')
            periodicidad_calibracion = row[5] if row[5] else 0
            fecha_calibracion = datetime.strptime(row[6], '%Y/%m/%d') if row[6] else None
            vencimiento_calibracion = datetime.strptime(row[7], '%Y/%m/%d') if row[7] else None

            # Verificar si el cod_articulo existe en la tabla indexssalud y obtiene las fechas actuales antes de actualizar
            cur = db.connection.cursor()
            cur.execute('SELECT estado_equipo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion FROM indexssalud WHERE cod_articulo = %s', [cod_articulo])
            producto = cur.fetchone()

            if not producto:
                # Si el cod_articulo no existe
                flash(f'El equipo con código {cod_articulo} no existe', 'error')
                continue  # Saltar a la siguiente fila del archivo

            estado_equipo = producto[0]
            fecha_mantenimiento_actual = producto[1]
            vencimiento_mantenimiento_actual = producto[2]
            fecha_calibracion_actual = producto[3]
            vencimiento_calibracion_actual = producto[4]

            if estado_equipo == "DE BAJA":
                flash(f'Equipo con código {cod_articulo} está dado de baja. No se actualizan fechas.', 'warning')
                continue

            # Verificar si alguna de las fechas ingresadas es diferente de las almacenadas
            if (fecha_mantenimiento.date() != fecha_mantenimiento_actual or 
                vencimiento_mantenimiento.date() != vencimiento_mantenimiento_actual or 
                fecha_calibracion != fecha_calibracion_actual or 
                vencimiento_calibracion != vencimiento_calibracion_actual):

                # Actualizar las fechas en la tabla indexssalud
                cur = db.connection.cursor()
                cur.execute("""UPDATE indexssalud SET periodicidad = %s, nombre_equipo = %s, fecha_mantenimiento = %s, vencimiento_mantenimiento = %s, periodicidad_calibracion = %s, fecha_calibracion = %s, vencimiento_calibracion = %s WHERE cod_articulo = %s""",
                                                     (periodicidad, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, periodicidad_calibracion, fecha_calibracion, vencimiento_calibracion, cod_articulo))
                db.connection.commit()
        
                # # Solo inserta en el historial si hubo cambios en las fechas de matenimiento y calibración
                # cur.execute("""INSERT INTO historial_fechas (cod_articulo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion, periodicidad, periodicidad_calibracion) VALUES (%s, %s, %s, %s, %s, %s, %s)""", 
                #                                             (cod_articulo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion, periodicidad, periodicidad_calibracion))
                # db.connection.commit()

                # Obtener hora actual del equipo
                hora_actual = datetime.now()
                
                # PARA EL CHECKBOX DE MANTENIMIENTO
                # checkbox_mantenimiento = 'Inactivo' # Valor predeterminado

                if hora_actual > vencimiento_mantenimiento:
                    color = 'purple'  # Vencido
                    # checkbox_mantenimiento = 'Inactivo'
                elif vencimiento_mantenimiento <= hora_actual + timedelta(days=30):
                    color = 'red'  # Falta menos de un mes
                    # checkbox_mantenimiento = 'Inactivo'
                elif vencimiento_mantenimiento <= hora_actual + timedelta(days=90):
                    color = 'yellow'  # Falta menos de tres meses
                    # checkbox_mantenimiento = 'Activo'
                else:
                    color = 'green'  # Vigente
                    # checkbox_mantenimiento = 'Activo' 
                
                # PARA EL CHECKBOX DE CALIBRACIÓN
                # checkbox_calibracion = 'Inactivo' # Valor predeterminado

                # if vencimiento_calibracion <= hora_actual + timedelta(days=30):
                #     checkbox_calibracion = 'Inactivo'
                # else:
                #     checkbox_calibracion = 'Activo'

                # Actualizar el color y los checkboxes en la tabla indexssalud
                cur.execute(""" UPDATE indexssalud SET color = %s WHERE cod_articulo = %s""", 
                                                      (color, cod_articulo))
                db.connection.commit()

            else:
                flash(f'Las fechas para el equipo {cod_articulo} no han cambiado', 'info')

        flash('Fechas actualizadas con éxito', 'success')
        return redirect(url_for('index_modulo', modulo='modulo'))
# ---------------------------FINALIZA ACTUALIZACIÓN MASIVA DE FECHAS CSV DE SALUD-----------------------------

# ---------------------------INICIA EXPORTACIÓN DE CSV DE EQUIPOS DE SALUD-----------------------------
@bp.route('/exportCsvsalud')
@login_required
def exportCsv():
    # Obtener los datos de la tabla indexssalud
    cur = db.connection.cursor()
    cur.execute("""SELECT i.cod_articulo, i.nombre_equipo, i.periodicidad, i.fecha_mantenimiento, i.vencimiento_mantenimiento, i.periodicidad_calibracion, 
                    i.fecha_calibracion, i.vencimiento_calibracion, i.fecha_ingreso, i.estado_equipo, i.ubicacion_original, i.garantia, i.criticos, 
                    p.nombre_empresa AS proveedor_responsable, i.checkbox_mantenimiento, i.checkbox_calibracion, i.especificaciones_instalacion,
                    i.cuidados_basicos, i.marca_equipo_salud, i.modelo_equipo_salud, i.serial_equipo_salud FROM indexssalud i LEFT JOIN datosproveedorsalud p ON i.proveedor_responsable = p.id
                    WHERE i.enable = 1""")
    registros = cur.fetchall()

    # Crear un archivo CSV en memoria
    si = StringIO()
    writer = csv.writer(si)

    # Escribir los encabezados de las columnas
    writer.writerow(['Código articulo', 'Nombre Equipo', 'Periodicidad Mantenimiento', 'Inicio Mantenimiento', 'Vencimiento Mantenimiento', 'Periodicidad Calibración', 'Inicio Calibración', 'Vencimiento Calibración', 'Fecha Ingreso', 
                      'Estado Equipo', 'Ubicación Original', 'Garantía Ingreso', 'Críticos', 'Proveedor Responsable', 'Check Mantenimiento', 'Check Calibración', 'Especificaciones Instalación', 'Cuidados Básicos', ' Marca Equipo', 'Modelo Equipo', 'Serial Equipo'])

    # Escribir los registros de la tabla
    for registro in registros:
        writer.writerow(registro)

    # Preparar el archivo CSV para su descarga
    salida = Response(si.getvalue().encode('utf-8-sig'), mimetype='text/csv')
    salida.headers['Content-Disposition'] = 'attachment; filename=equiposSalud.csv'

    return salida
# ---------------------------FINALIZA EXPORTACIÓN DE CSV DE EQUIPOS DE SALUD-----------------------------

# ---------------------------INICIA EXPORTACIÓN DE FORMATO EXCEL DE EQUIPOS DE BAJA DE SALUD-----------------------------
@bp.route('/exportExcelSaludDeBaja', methods=['POST'])
@login_required
def exportExcelDeBaja():
    try:
        # 1) Validar JSON/csrf
        if not request.is_json:
            return jsonify({"error": "El cuerpo de la petición debe ser JSON"}), 400

        data = request.get_json(silent=True) or {}
        equipos = data.get('equipos') or []

        if not isinstance(equipos, list) or not equipos:
            return jsonify({"error": "No se enviaron equipos"}), 400

        # 2) Ubicar plantilla
        plantilla_path = os.path.join(current_app.root_path, "static", "img", "INFORME_TECNICO_BAJAS.xlsx")
        if not os.path.exists(plantilla_path):
            return jsonify({"error": f"No se encontró la plantilla en {plantilla_path}"}), 400

        # 3) Cargar plantilla
        wb = load_workbook(plantilla_path)
        ws = wb.active  # Ajusta a wb['NombreHoja'] si tu hoja no es la activa

        # 4) Escribir datos desde fila 12 (debajo de los títulos)
        start_row = 12

        # Helper: verificar si un rango ya está fusionado
        def is_merged(ws, coord: str) -> bool:
            return any(str(r) == coord for r in ws.merged_cells.ranges)

        for idx, equipo in enumerate(equipos, start=start_row):
            placa = (equipo.get("cod_articulo") or "").strip()
            cantidad = 1
            nombre = (equipo.get("nombre_equipo") or "").strip()

            # Columna A → PLACA
            ws[f"A{idx}"].value = placa
                       
            # Columna B → CANTIDAD
            ws[f"B{idx}"].value = cantidad

            # Columnas C-G → DESCRIPCIÓN DEL BIEN (escribir en C, y fusionar sólo si hace falta)
            merge_coord = f"C{idx}:G{idx}"
            if not is_merged(ws, merge_coord):
                try:
                    ws.merge_cells(merge_coord)
                except ValueError:
                    # Si ya está fusionado por plantilla o solapa, lo ignoramos
                    pass

            ccell = ws[f"C{idx}"]
            ccell.value = nombre
            ccell.alignment = Alignment(wrap_text=True, vertical="top")

        # 5) Enviar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="equipos_baja.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        current_app.logger.exception("Error generando Excel")
        return jsonify({"error": str(e)}), 400
# ---------------------------FINALIZA EXPORTACIÓN DE CSV DE EQUIPOS DE EQUIPOS DE BAJA DE SALUD-----------------------------

# ================================CHECKBOX PROGRAMACIÓN MANTENIMIENTO===============================
@bp.route('/checkbox_programacionMantenimiento', methods=['POST'])
def checkbox_programacionMantenimiento():
    try:
        seleccionados = request.form.getlist('seleccionados[]')
        proveedor_id = request.form.get('proveedor_id')

        if not seleccionados or not proveedor_id:
            return jsonify({'success': False, 'message': 'Faltan productos seleccionados o proveedor.'})

        hoy = datetime.now()
        cur = db.connection.cursor()

        productos_guardados = []

        for cod in seleccionados:
            nombre_equipo = request.form.get(f'nombre_equipo_{cod}')
            ubicacion = request.form.get(f'ubicacion_{cod}')
            periodicidad_m = request.form.get(f'periodicidad_mantenimiento_{cod}')
            periodicidad_c = request.form.get(f'periodicidad_calibracion_{cod}')

            mantenimiento_activado = request.form.get(f'mantenimiento_{cod}') == 'on'
            calibracion_activada = request.form.get(f'calibracion_{cod}') == 'on'

            # Obtener fechas reales desde base de datos
            cur.execute("SELECT fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion FROM indexssalud WHERE cod_articulo = %s", (cod,))
            resultado = cur.fetchone()
            if not resultado:
                continue  # Saltar si no se encuentra

            fecha_m, vencimiento_m, fecha_c, vencimiento_c = resultado

            if mantenimiento_activado:
                # Validación de vencimiento
                if vencimiento_m and (vencimiento_m - hoy).days < 30:
                    continue
                # Activar checkbox y guardar en historial
                cur.execute(
                    "UPDATE indexssalud SET checkbox_mantenimiento = 'Activo' WHERE cod_articulo = %s",
                    (cod,),
                )
                cur.execute(
                    """INSERT INTO historial_mantenimiento_salud 
                            (cod_articulo, nombre_equipo, ubicacion_original, fecha_mantenimiento, vencimiento_mantenimiento, periodicidad, id_proveedor_responsable) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                    (
                        cod,
                        nombre_equipo,
                        ubicacion,
                        fecha_m,
                        vencimiento_m,
                        periodicidad_m,
                        proveedor_id,
                    ),
                )

            if calibracion_activada:
                if vencimiento_c and (vencimiento_c - hoy).days < 30:
                    continue
                cur.execute(
                    "UPDATE indexssalud SET checkbox_calibracion = 'Activo' WHERE cod_articulo = %s",
                    (cod,),
                )
                cur.execute(
                    """INSERT INTO historial_calibracion_salud 
                            (cod_articulo, nombre_equipo, ubicacion_original, fecha_calibracion, vencimiento_calibracion, periodicidad_calibracion, id_proveedor_responsable) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                    (
                        cod,
                        nombre_equipo,
                        ubicacion,
                        fecha_c,
                        vencimiento_c,
                        periodicidad_c,
                        proveedor_id,
                    ),
                )

            productos_guardados.append(cod)

        db.connection.commit()
        return jsonify({
            'success': True,
            'message': f'Se procesaron {len(productos_guardados)} productos correctamente.',
            'productos': productos_guardados
        })

    except Exception as e:
        db.connection.rollback()
        return jsonify({'success': False, 'message': f'Error en el servidor: {str(e)}'})
    
# ======================================================================================================

@bp.route('/guardar_historial_masivo', methods=['POST'])
@login_required
def guardar_historial_masivo():
    data = request.get_json()
    proveedor_id = data.get('proveedorId')
    nueva_fecha_str = data.get('nuevaFecha')
    registros = data.get('registros', [])

    try:
        if not proveedor_id or not nueva_fecha_str:
            return jsonify({'success': False, 'message': 'Falta proveedor o fecha'})

        nueva_fecha = datetime.strptime(nueva_fecha_str, '%Y-%m-%d')
        cur = db.connection.cursor()

        for r in registros:
            tipo = r.get('tipo')  # fecha_mantenimiento o fecha_calibracion
            producto_id = r.get('productoId')
            nueva_periodicidad = int(data.get('nuevaPeriodicidad', 0))
            nombre_equipo = r.get('nombreEquipo')
            ubicacion = r.get('ubicacionOriginal')

            # Obtener datos actuales para historial mantenimiento
            if tipo == "fecha_mantenimiento":
                cur.execute("SELECT fecha_mantenimiento, vencimiento_mantenimiento, periodicidad FROM indexssalud WHERE cod_articulo = %s", (producto_id,))
                resultado = cur.fetchone()
                if not resultado:
                    continue
                fecha_actual, vencimiento_actual, periodicidad_actual = resultado

                # Guardar en historial
                cur.execute(
                    """INSERT INTO historial_mantenimiento_salud 
                    (cod_articulo, nombre_equipo, ubicacion_original, fecha_mantenimiento, vencimiento_mantenimiento, periodicidad, id_proveedor_responsable)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                    (
                        producto_id,
                        nombre_equipo,
                        ubicacion,
                        fecha_actual,            # fecha anterior guardada
                        nueva_fecha,             # vencimiento será la nueva fecha elegida
                        periodicidad_actual,
                        proveedor_id,
                    )
                )

                # Calcular nuevo vencimiento
                nuevo_vencimiento = nueva_fecha + relativedelta(months=nueva_periodicidad)

                # Actualizar en indexssalud
                cur.execute(
                    "UPDATE indexssalud SET fecha_mantenimiento = %s, vencimiento_mantenimiento = %s, periodicidad = %s WHERE cod_articulo = %s",
                    (nueva_fecha, nuevo_vencimiento, nueva_periodicidad, producto_id)
                )
            
            # Obtener datos actuales para historial calibracion
            elif tipo == "fecha_calibracion":
                cur.execute("SELECT fecha_calibracion, vencimiento_calibracion, periodicidad_calibracion FROM indexssalud WHERE cod_articulo = %s", (producto_id,))
                resultado = cur.fetchone()
                if not resultado:
                    continue
                fecha_actual, vencimiento_actual, periodicidad_actual = resultado

                # Guardar en historial
                cur.execute(
                    """INSERT INTO historial_calibracion_salud 
                    (cod_articulo, nombre_equipo, ubicacion_original, fecha_calibracion, vencimiento_calibracion, periodicidad_calibracion, id_proveedor_responsable)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                    (
                        producto_id,
                        nombre_equipo,
                        ubicacion,
                        fecha_actual,
                        nueva_fecha,
                        periodicidad_actual,
                        proveedor_id,
                    )
                )

                # Calcular nuevo vencimiento
                nuevo_vencimiento = nueva_fecha + relativedelta(months=nueva_periodicidad)

                # Actualizar en indexssalud
                cur.execute(
                    "UPDATE indexssalud SET fecha_calibracion = %s, vencimiento_calibracion = %s, periodicidad_calibracion = %s WHERE cod_articulo = %s",
                    (nueva_fecha, nuevo_vencimiento, nueva_periodicidad, producto_id)
                )

        db.connection.commit()
        return jsonify({'success': True, 'message': 'Fechas y registros actualizados correctamente.'})

    except Exception as e:
        db.connection.rollback()
        print("Error:", e)
        return jsonify({'success': False, 'message': f'Error en el servidor: {str(e)}'})

# ================================CHECKBOX PROGRAMACIÓN FECHA DE CALIBRACIÓN===============================
# @app.route('/checkbox_programacionCalibracion', methods=['POST'])
# def checkbox_programacionCalibracion():
#     if request.method == 'POST':
#             data = request.get_json()
#             producto_id = data['productoId']
#             nuevo_estado = data['nuevoEstado']
#             # checkbox_mantenimiento = data.get('CheckboxMantenimiento')

#             cur = db.connection.cursor()
#             cur.execute('UPDATE indexssalud SET checkbox_calibracion = %s WHERE id = %s', (nuevo_estado, producto_id))
#             db.connection.commit()

#             return redirect(url_for('indexSalud'))

#     return jsonify({'error': 'Método no permitido'}), 405
# =====================================================================================================

# ACTUALIZA EL ESTADO DEL EQUIPO DESDE EL DESPLEGABLE QUE SE ENCUENTRA EN LA MISMA TABLA INDEXSALUD
@bp.route('/<modulo>/update_estado_equipo', methods=['POST'])
def update_estado_equipo(modulo):
    if request.method == 'POST':

        # OBTENER FULLNAME DEL USUARIO LOGUEADO
        cur = db.connection.cursor()
        cur.execute("SELECT fullname, username FROM user WHERE id = %s", (current_user.id,))
        result = cur.fetchone()
        usuario_logueado_nombre = result[0] if result else None
        usuario_logueado_email  = result[1] if result else None

        quien_recibe_equipo = request.form.get('quien_recibe_equipo')
        email_receptor = request.form.get('email_receptor')
        ubicacion_destino_laboratorio = request.form.get('ubicacion_destino_laboratorio')

        producto_id = request.form['producto_id']
        nuevo_estado = request.form['nuevo_estado_equipo']
        cod_articulo = request.form ['cod_articulo']
        nombre_equipo = request.form ['nombre_equipo']

        # Obtener hora actual del equipo
        hora_actual = datetime.now()

        # PARA EL CHECKBOX Y SEMAFORO DE MANTENIMIENTO
        fecha_mantenimiento = request.form ['fecha_mantenimiento']
        vencimiento_mantenimiento = request.form ['vencimiento_mantenimiento']

        if fecha_mantenimiento and vencimiento_mantenimiento:
            fecha_mantenimiento = datetime.strptime(fecha_mantenimiento, '%Y-%m-%d')
            vencimiento_mantenimiento = datetime.strptime(vencimiento_mantenimiento, '%Y-%m-%d')

            if hora_actual > vencimiento_mantenimiento:
                color = 'purple'  # Vencido
            elif vencimiento_mantenimiento <= hora_actual + timedelta(days=30):
                color = 'red'  # Falta menos de un mes
            elif vencimiento_mantenimiento <= hora_actual + timedelta(days=90):
                color = 'yellow'  # Falta menos de tres meses
            else:
                color = 'green'  # Vigente
        else:
            flash('Debe ingresar las fechas de inicio y vencimiento de mantenimiento.', 'error')
            return redirect(url_for('index_modulo', modulo=modulo))
        
        # PARA EL CHECKBOX DE CALIBRACIÓN
        fecha_calibracion = request.form.get('fecha_calibracion', '') or None
        vencimiento_calibracion = request.form.get('vencimiento_calibracion', '') or None
       
        # Convertir a objetos date solo si existen
        if fecha_calibracion:
            try:
                fecha_calibracion = datetime.strptime(fecha_calibracion, '%Y-%m-%d').date()
            except ValueError:
                fecha_calibracion = None

        if vencimiento_calibracion:
            try:
                vencimiento_calibracion = datetime.strptime(vencimiento_calibracion, '%Y-%m-%d').date()
            except ValueError:
                vencimiento_calibracion = None

        fecha_ingreso = request.form ['fecha_ingreso']
        periodicidad = request.form ['periodicidad']
        estado_equipo= request.form ['estado_equipo']
        ubicacion_original = request.form ['ubicacion_original']
        garantia = request.form ['garantia']
        criticos = request.form ['criticos']
        proveedor_responsable = request.form ['proveedor_responsable']
        especificaciones_instalacion = request.form ['especificaciones_instalacion']
        cuidados_basicos = request.form ['cuidados_basicos']
        periodicidad_calibracion = request.form ['periodicidad_calibracion']
        marca_equipo_salud = request.form ['marca_equipo_salud']
        modelo_equipo_salud = request.form ['modelo_equipo_salud']
        serial_equipo_salud = request.form ['serial_equipo_salud']
        cur = db.connection.cursor()

        # Obtener la ruta de la imagen desde la tabla indexssalud
        cur.execute('SELECT imagen FROM indexssalud WHERE cod_articulo = %s', (cod_articulo,))
        imagen_result = cur.fetchone()
        filepath_to_db_img = imagen_result[0] if imagen_result else None

        if nuevo_estado == 'DE BAJA':
            # Actualizar el estado y marcar como dado de baja en indexssalud
            cur.execute("""UPDATE indexssalud SET estado_equipo = %s, enable = 0, de_baja = 1 WHERE cod_articulo = %s""", (nuevo_estado, cod_articulo))

            # Verificar si el equipo ya está en equipossalud_debaja
            cur.execute('SELECT * FROM equipossalud_debaja WHERE cod_articulo = %s', (cod_articulo,))
            equipo_existente = cur.fetchone()

            # Insertar el equipo en equipossalud_debaja si no existe
            if not equipo_existente:
                cur.execute("""INSERT INTO equipossalud_debaja (cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion, fecha_ingreso,
                                                                periodicidad, estado_equipo, ubicacion_original, garantia, criticos, proveedor_responsable, color, imagen, especificaciones_instalacion, cuidados_basicos,
                                                                periodicidad_calibracion, marca_equipo_salud, modelo_equipo_salud, serial_equipo_salud, fecha_de_baja) 
                                                                VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (
                        cod_articulo,
                        nombre_equipo,
                        fecha_mantenimiento,
                        vencimiento_mantenimiento,
                        fecha_calibracion,
                        vencimiento_calibracion,
                        fecha_ingreso,
                        periodicidad,
                        nuevo_estado,  # Estado del equipo
                        ubicacion_original,
                        garantia,
                        criticos,
                        proveedor_responsable,
                        color,
                        filepath_to_db_img,
                        especificaciones_instalacion,
                        cuidados_basicos,
                        periodicidad_calibracion,
                        marca_equipo_salud,
                        modelo_equipo_salud,
                        serial_equipo_salud,
                        hora_actual
                    ),
                )
        else:
            # Actualiza el estado si no está dado de baja
            cur.execute('UPDATE indexssalud SET estado_equipo = %s WHERE cod_articulo = %s', (nuevo_estado, cod_articulo))

        fecha_prestamo = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # --- 1. Si se está devolviendo un equipo ---
        if nuevo_estado != 'PRÉSTAMO':
            cur.execute("""
                SELECT id_prestamos, quien_recibe_equipo, email_receptor
                FROM prestamos_equiposalud
                WHERE cod_articulo = %s AND enable_prestamos = 1
                ORDER BY id_prestamos DESC
                LIMIT 1
            """, (cod_articulo,))
            prestamo_activo = cur.fetchone()

            # Cerrar préstamo y registrar quien lo recibe devuelto
            cur.execute("""
                UPDATE prestamos_equiposalud
                SET enable_prestamos = 0,
                    fecha_entrega_equipo = %s,
                    quien_recibe_equipo_devuelto = %s
                WHERE cod_articulo = %s AND enable_prestamos = 1
            """, (fecha_prestamo, usuario_logueado_nombre, cod_articulo))
            cerrados = cur.rowcount

            if prestamo_activo and cerrados > 0:
                send_devolucion_notification_html(
                    equipo_nombre=nombre_equipo,
                    codigo_equipo=cod_articulo,
                    quien_devuelve=prestamo_activo[1],        # Quien recibió el préstamo originalmente
                    quien_recibe=usuario_logueado_nombre,     # Usuario logueado recibe la devolución
                    email_entrega=usuario_logueado_email,     # Email usuario logueado
                    email_recibe=prestamo_activo[2],          # Email receptor original
                    fecha_entrega=fecha_prestamo
                )

        # --- 2. Actualizar el estado del equipo ---
        cur.execute("""
            UPDATE indexssalud
            SET estado_equipo = %s
            WHERE cod_articulo = %s
        """, (nuevo_estado, cod_articulo))

        # --- 3. Si es un nuevo préstamo ---
        if nuevo_estado == 'PRÉSTAMO':
            ubicacion_original = request.form.get('ubicacion_original')
            cur.execute("""
                INSERT INTO prestamos_equiposalud (
                    cod_articulo, nombre_equipo, usuario_logueado_nombre, quien_recibe_equipo,
                    ubicacion_original, ubicacion_destino_laboratorio, fecha_prestamo_equipo,
                    fecha_entrega_equipo, email_receptor, quien_recibe_equipo_devuelto, enable_prestamos
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, NULL, %s, NULL, 1)
            """, (
                cod_articulo,
                nombre_equipo,
                usuario_logueado_nombre,
                quien_recibe_equipo,
                ubicacion_original,
                ubicacion_destino_laboratorio,
                fecha_prestamo,
                email_receptor
            ))

            send_prestamo_notification_html(
                cod_articulo,
                nombre_equipo,
                usuario_logueado_nombre,
                ubicacion_original,
                usuario_logueado_email,
                quien_recibe_equipo,
                ubicacion_destino_laboratorio,
                email_receptor,
                fecha_prestamo
            )

        db.connection.commit()
        cur.close()
        flash('Estado del equipo actualizado correctamente', 'success')
        return redirect(url_for('inventario.index_modulo', modulo='modulo'))
# =======================================//=======================================================
# FUNCIÓN PARA MOSTRAR LOS DATOS DE LOS EQUIPOS QUE SE ENCUENTRAN PRESTADOS
@bp.route('/prestamos_equipos_salud/<cod_articulo>')
@login_required
def prestamos_equipos_salud(cod_articulo):
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

    # Consultar los datos del prestamo en la tabla prestamos_equiposalud
    cur.execute('SELECT * FROM prestamos_equiposalud WHERE cod_articulo = %s AND enable_prestamos = 1 ORDER BY id_prestamos DESC LIMIT 1', (cod_articulo,))
    prestamo_equipo = cur.fetchone()

    if not prestamo_equipo:
        flash('No se encontraron datos para este equipo.', 'warning')
        return redirect(url_for('inventario.index_modulo'))

    return render_template('prestamos_equipos_salud.html', prestamo_equipo=prestamo_equipo)
# =====================================================================================================

# =======================================FUNCIÓN PARA HISTORIAL DE PRESTAMOS DE SALUD==============================
@bp.route('/historialPrestamoSalud')
@login_required
def historial_prestamo_salud():
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

    try:
        cur.execute(
            """
            SELECT cod_articulo, nombre_equipo, ubicacion_original, ubicacion_destino_laboratorio, 
                   quien_recibe_equipo, usuario_logueado_nombre, fecha_prestamo_equipo, fecha_entrega_equipo, email_receptor, quien_recibe_equipo_devuelto, enable_prestamos
            FROM prestamos_equiposalud
            ORDER BY cod_articulo ASC
        """,   
        )

        historial = cur.fetchall()
        return render_template('historialPrestamoSalud.html', prestamos_equiposalud=historial)

    except Exception as e:
        print(f"Error al obtener el historial: {str(e)}")
        flash('Error al obtener el historial de prestamos.', 'danger')
        return redirect(url_for('index_modulo', modulo='modulo'))
    finally:
        cur.close()
# =====================================================================================================

# ESTA FUNCIÓN ME LLEVA A OTRA VISTA TRAYENDO LOS PARAMETROS DE AGREGAR PARA DESPUES PODER ACTUALIZAR EN LA SIGUIENTE FUNCIÓN. LAS DOS SE COMPLEMENTAN
@bp.route('/edit_productoSalud/<id>/<vista>', methods=['GET'])
@login_required
def GET_PRODUCTO_SALUD(id, vista):
    cur = db.connection.cursor()

    # Verificar de qué vista proviene la solicitud para seleccionar la tabla correcta
    if vista == 'index_modulo':
        cur.execute('SELECT * FROM indexssalud WHERE id = %s', [id])
    elif vista == 'equiposDeBajaSalud':
        cur.execute('SELECT * FROM equipossalud_debaja WHERE id = %s', [id])

    data = cur.fetchall()

    if data:
        cod_articulo = data[0][1]
    print(cod_articulo)    
    # Obtener el historial de fechas mantenimiento del equipo
    cur.execute("""SELECT * FROM historial_mantenimiento_salud WHERE cod_articulo = %s ORDER BY fecha_mantenimiento DESC""", [cod_articulo])
    historial_mantenimiento = cur.fetchall()

    # Obtener el historial de fechas calibracion del equipo
    cur.execute("""SELECT * FROM historial_calibracion_salud WHERE cod_articulo = %s ORDER BY fecha_calibracion DESC""", [cod_articulo])
    historial_calibracion = cur.fetchall()

    # Combinar historial para enviar al template (si lo necesitas despues)
    historial = {
        'mantenimiento': historial_mantenimiento,
        'calibracion': historial_calibracion
    }
    
    return render_template('editar_producto_indexSalud.html', producto = data[0], cod_articulo = cod_articulo, historial = historial)

# FUNCIÓN ACTUALIZAR EDITAR/VER HOJA DE VIDA
@bp.route('/actualizarSalud/<id>', methods = ['POST'])
def ACTUALIZAR_PRODUCTO_SALUD(id):
    if request.method =='POST':
        cod_articulo = request.form ['cod_articulo']
        nombre_equipo = request.form ['nombre_equipo']
        
        # PARA EL CHECKBOX Y SEMAFORO DE MANTENIMIENTO
        fecha_mantenimiento = request.form ['fecha_mantenimiento']
        vencimiento_mantenimiento = request.form ['vencimiento_mantenimiento']
        
        # Obtener hora actual del equipo
        hora_actual = datetime.now().date()
        color = 'verde'

        if  vencimiento_mantenimiento:
            vencimiento_mant = datetime.strptime(vencimiento_mantenimiento, '%Y-%m-%d').date()
            if vencimiento_mant < hora_actual + timedelta(days=0):
                color = 'purple'  # Falta menos de un mes
            elif vencimiento_mant <= hora_actual + timedelta(days=30):
                color = 'red'  # Falta menos de tres meses
            elif vencimiento_mant <= hora_actual + timedelta(days=90):
                color = 'yellow'  # Falta menos de tres meses
        else:
            flash('Debe ingresar las fechas de mantenimiento.', 'error')
            return redirect(url_for('inventario.index_modulo'))
        
        fecha_calibracion = request.form.get('fecha_calibracion') or None
        vencimiento_calibracion = request.form.get('vencimiento_calibracion') or None

        fecha_ingreso = request.form ['fecha_ingreso']
        periodicidad = request.form ['periodicidad']
        periodicidad_calibracion = request.form ['periodicidad_calibracion']
        garantia = request.form ['garantia']
        especificaciones_instalacion = request.form ['especificaciones_instalacion']
        cuidados_basicos = request.form ['cuidados_basicos']
        
        cur = db.connection.cursor() 

        # Obtener las fechas actuales antes de actualizar
        cur.execute(
            """ UPDATE indexssalud SET cod_articulo = %s, nombre_equipo = %s, fecha_mantenimiento = %s, vencimiento_mantenimiento = %s, fecha_calibracion = %s, vencimiento_calibracion = %s, 
                fecha_ingreso = %s, periodicidad = %s, garantia = %s, color = %s, especificaciones_instalacion = %s, cuidados_basicos = %s, periodicidad_calibracion = %s WHERE id = %s""",
            (
                cod_articulo,
                nombre_equipo,
                fecha_mantenimiento,
                vencimiento_mantenimiento,
                fecha_calibracion,
                vencimiento_calibracion,
                fecha_ingreso,
                periodicidad,
                garantia,
                color,
                especificaciones_instalacion,
                cuidados_basicos,
                periodicidad_calibracion,
                id,
            ),
        )
        db.connection.commit()

        flash('Equipo actualizado satisfactoriamente', 'success')
        return redirect(url_for('inventario.index_modulo', modulo='modulo')) 

# HISTORIAL FECHAS MANTENIMIENTO Y CALIBRACIÓN
@bp.route('/historialFechas/<cod_articulo>')
@login_required
def historialFechas(cod_articulo):
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    try:
        cur.execute(
            """
            SELECT id, cod_articulo, nombre_equipo, ubicacion_original, fecha_mantenimiento, 
                   vencimiento_mantenimiento, periodicidad, id_proveedor_responsable
            FROM historial_mantenimiento_salud 
            WHERE cod_articulo = %s 
            ORDER BY fecha_mantenimiento DESC
        """,
            [cod_articulo],
        )
        mantenimiento = cur.fetchall()

        cur.execute(
            """
            SELECT id, cod_articulo, nombre_equipo, ubicacion_original, fecha_calibracion, 
                   vencimiento_calibracion, periodicidad_calibracion, id_proveedor_responsable
            FROM historial_calibracion_salud 
            WHERE cod_articulo = %s 
            ORDER BY fecha_calibracion DESC
        """,
            [cod_articulo],
        )
        calibracion = cur.fetchall()

        historial = {
            'mantenimiento': mantenimiento,
            'calibracion': calibracion
        }

        return render_template('historialFechas.html', historial=historial)

    except Exception as e:
        print(f"Error al obtener el historial: {str(e)}")
        flash('Error al obtener el historial de fechas.', 'danger')
        return redirect(url_for('index_modulo', modulo='modulo'))
    finally:
        cur.close()


# ACTUALIZAR FECHAS DE MANTENIMIENTO
@bp.route('/update_historial_fechas', methods=['POST'])
def update_historial_fechas():
    id = request.form['id']  # ID del registro en historial_fechas
    cod_articulo = request.form['cod_articulo']
    fecha_mantenimiento = request.form['fecha_mantenimiento']
    vencimiento_mantenimiento = request.form['vencimiento_mantenimiento']
    periodicidad = request.form['periodicidad']

    # Usar DictCursor para obtener un diccionario
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

    # Verificar si el registro es el último de su tipo en historial_fechas
    cur.execute("""SELECT id FROM historial_mantenimiento_salud WHERE cod_articulo = %s ORDER BY id DESC LIMIT 1""", 
                [cod_articulo])
    last_record = cur.fetchone()

   # Si el id del último registro coincide con el seleccionado, permitir la actualización
    if last_record and last_record['id'] == int(id):
        # Actualizar en historial_fechas
        cur = db.connection.cursor()
        cur.execute("""UPDATE historial_mantenimiento_salud SET fecha_mantenimiento = %s, 
                       vencimiento_mantenimiento = %s, periodicidad = %s WHERE id = %s""", 
                      (fecha_mantenimiento, vencimiento_mantenimiento, periodicidad, id))
        
        # Determinar el color del semáforo basado en la fecha de vencimiento más próxima
        fecha_actual = datetime.now().date()
        color = "verde"  # Valor por defecto

        # Comparar fechas de mantenimiento y calibración
        if vencimiento_mantenimiento:
            
            vencimiento_mant = datetime.strptime(vencimiento_mantenimiento, '%Y-%m-%d').date()
            if vencimiento_mant < fecha_actual + timedelta(days=0):
                color = "purple"
            elif vencimiento_mant <= fecha_actual + timedelta(days=30):
                color = "red"
            elif vencimiento_mant <= fecha_actual + timedelta(days=90):
                color = "yellow"
        
        db.connection.commit()
        flash('Historial actualizado correctamente', 'success')
    else:
        # Si no es el último registro, mostrar un mensaje de error
        flash('Solo se puede actualizar el último registro de este equipo.', 'danger')
    
    return redirect(url_for('historialFechas', cod_articulo=cod_articulo))

# ACTUALIZAR FECHAS DE HISTORIAL CALIBRACIÓN
@bp.route('/update_historial_fechas_calibracion', methods=['POST'])
def update_historial_fechas_calibracion():
    id = request.form['id']  # ID del registro en historial_fechas_calibracion
    cod_articulo = request.form['cod_articulo']
    fecha_calibracion = request.form['fecha_calibracion']
    vencimiento_calibracion = request.form['vencimiento_calibracion']
    periodicidad_calibracion = request.form['periodicidad_calibracion']

    # Usar DictCursor para obtener un diccionario
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

    # Verificar si el registro es el último de su tipo en historial_fechas
    cur.execute("""SELECT id FROM historial_calibracion_salud WHERE cod_articulo = %s ORDER BY id DESC LIMIT 1 """, 
                   [cod_articulo])
    last_record = cur.fetchone()

   # Si el id del último registro coincide con el seleccionado, permitir la actualización
    if last_record and last_record['id'] == int(id):
        # Actualizar en historial_fechas
        cur = db.connection.cursor()
        cur.execute(""" UPDATE historial_calibracion_salud SET fecha_calibracion = %s, 
                        vencimiento_calibracion = %s, periodicidad_calibracion = %s WHERE id = %s""", 
                        (fecha_calibracion, vencimiento_calibracion, periodicidad_calibracion, id))
        
        db.connection.commit()
        flash('Historial actualizado correctamente', 'success')
    else:
        # Si no es el último registro, mostrar un mensaje de error
        flash('Solo se puede actualizar el último registro de este equipo.', 'danger')
    
    return redirect(url_for('historialFechas', cod_articulo=cod_articulo))


# ==========================INICIA FUNCIÓN EQUIPOS DADOS DE BAJA=====================
@bp.route('/equiposDeBajaSalud')
@login_required
def equiposDeBajaSalud():
    
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute('SELECT id, estado_equipo FROM estados_equipos')
    estadoEquipos = cur.fetchall()
    
    cur.execute('SELECT * FROM equipossalud_debaja')
    equipos_de_baja = cur.fetchall()
    return render_template('equiposDeBajaSalud.html', equipos_de_baja=equipos_de_baja, estadoEquipos=estadoEquipos)
# ==========================FINALIZA FUNCIÓN EQUIPOS DADOS DE BAJA=====================


# FUNCIÓN ELIMINAR PARA INDEXSSALUD
@bp.route('/delete_productoSalud/<string:id>')
@login_required
def ELIMINAR_CONTACTO_SALUD(id):
    cur = db.connection.cursor()
    # cur.execute('DELETE FROM indexssalud WHERE id = {0}'.format(id))
    #Esta linea de codigo en la vista elimina el producto pero no de la DB, la cual realiza es una actualización
    cur.execute('UPDATE indexssalud SET enable=0 WHERE id = {0}'.format(id))
    db.connection.commit()
    flash('Equipo eliminado satisfactoriamente', 'success')
    return redirect(url_for('index_modulo', modulo='modulo'))